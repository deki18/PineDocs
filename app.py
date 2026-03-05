import os
import tempfile
import base64
import io
from typing import List, Tuple, Optional
from datetime import datetime
from pathlib import Path

import streamlit as st
from dotenv import load_dotenv
import requests
import fitz  # PyMuPDF

from pinecone import Pinecone, ServerlessSpec, CloudProvider, AwsRegion
from openai import OpenAI

load_dotenv()

PINECONE_API_KEY = os.getenv("PINECONE_API_KEY")
VECTORENGINE_API_KEY = os.getenv("VECTORENGINE_API_KEY")

DOCS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "docs")
os.makedirs(DOCS_DIR, exist_ok=True)


def get_files_from_folder(folder_path: str, include_subfolders: bool = True) -> List[Path]:
    """递归获取文件夹中的所有文件
    
    Args:
        folder_path: 文件夹路径
        include_subfolders: 是否包含子文件夹
        
    Returns:
        文件路径列表
    """
    folder = Path(folder_path)
    if not folder.exists() or not folder.is_dir():
        return []
    
    supported_extensions = {'.txt', '.md', '.doc', '.docx', '.xlsx', '.pdf', '.jpg', '.jpeg', '.png', '.bmp'}
    files = []
    
    if include_subfolders:
        for file_path in folder.rglob('*'):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                files.append(file_path)
    else:
        for file_path in folder.glob('*'):
            if file_path.is_file() and file_path.suffix.lower() in supported_extensions:
                files.append(file_path)
    
    return sorted(files)


def create_file_object(file_path: Path):
    """创建类似UploadedFile的对象
    
    Args:
        file_path: 文件路径
        
    Returns:
        包含文件信息的对象
    """
    class FileObject:
        def __init__(self, path: Path):
            self.name = path.name
            self.path = str(path)
            self.size = path.stat().st_size
            self._path = path
        
        def read(self):
            with open(self._path, 'rb') as f:
                return f.read()
    
    return FileObject(file_path)


@st.cache_data(ttl=300)
def get_embed_client():
    if not VECTORENGINE_API_KEY:
        return None
    return OpenAI(
        base_url="https://api.vectorengine.ai/v1",
        api_key=VECTORENGINE_API_KEY,
    )


def get_pinecone_index(index_name: str):
    if not PINECONE_API_KEY:
        st.error("请在.env文件中设置PINECONE_API_KEY")
        return None

    pc = Pinecone(api_key=PINECONE_API_KEY)

    existing_indexes = [idx["name"] for idx in pc.list_indexes()]
    if index_name not in existing_indexes:
        pc.create_index(
            name=index_name,
            dimension=3072,
            metric="cosine",
            spec=ServerlessSpec(
                cloud=CloudProvider.AWS,
                region=AwsRegion.US_EAST_1
            ),
        )

    return pc.Index(index_name)


def embed_texts(texts: List[str]) -> List[List[float]]:
    client = get_embed_client()
    if not client:
        return []

    response = client.embeddings.create(
        model="text-embedding-3-large",
        input=texts,
    )

    return [item.embedding for item in response.data]


def split_text(text: str, max_chars: int = 1000, overlap: int = 200) -> List[str]:
    if max_chars <= overlap:
        raise ValueError("max_chars must be greater than overlap")

    chunks: List[str] = []
    start = 0
    length = len(text)

    while start < length:
        end = min(start + max_chars, length)
        
        if end < length:
            last_newline = text.rfind('\n', start, end)
            last_period = text.rfind('。', start, end)
            last_exclamation = text.rfind('！', start, end)
            last_question = text.rfind('？', start, end)
            last_semicolon = text.rfind('；', start, end)
            
            best_break = max(last_newline, last_period, last_exclamation, last_question, last_semicolon)
            
            if best_break > start + max_chars // 2:
                end = best_break + 1
        
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        
        if end == length:
            break
        
        start = max(end - overlap, 0)

    return chunks


def pdf_to_markdown_with_ollama(pdf_bytes: bytes, ollama_url: str = "http://localhost:11434", model_name: str = "my-glm-ocr:latest", filename: str = "document.pdf") -> str:
    """使用Ollama的OCR模型将PDF转换为Markdown
    
    支持的模型:
    - my-PaddleOCR-VL:0.9b: PaddleOCR-VL-1.5 (推荐，效果更好)
    - my-glm-ocr:latest: GLM-OCR模型
    
    Args:
        pdf_bytes: PDF文件的字节数据
        ollama_url: Ollama服务地址
        model_name: 使用的OCR模型名称
        filename: 原始文件名，用于保存Markdown文件
    """
    try:
        # 使用PyMuPDF将PDF转换为图片列表
        st.info("正在将PDF转换为图片...")
        try:
            # 打开PDF
            pdf_document = fitz.open(stream=pdf_bytes, filetype="pdf")
            images = []
            
            # 设置缩放比例 (zoom=2.0 相当于 144 DPI, zoom=3.0 相当于 216 DPI)
            zoom = 2.0
            mat = fitz.Matrix(zoom, zoom)
            
            for page_num in range(len(pdf_document)):
                page = pdf_document[page_num]
                # 将页面渲染为图片
                pix = page.get_pixmap(matrix=mat)
                # 转换为PIL Image
                img_data = pix.tobytes("png")
                images.append(img_data)
            
            pdf_document.close()
            
        except Exception as e:
            st.error(f"PDF转图片失败: {str(e)}")
            return ""
        
        if not images:
            st.error("PDF中没有找到页面")
            return ""
        
        st.info(f"PDF共 {len(images)} 页，正在逐页识别...")
        
        all_content = []
        
        # 逐页处理
        for page_num, img_data in enumerate(images, start=1):
            st.info(f"正在处理第 {page_num}/{len(images)} 页...")
            
            # 将图片转换为base64
            img_base64 = base64.b64encode(img_data).decode('utf-8')
            
            # 根据模型选择不同的调用方式
            if model_name == "my-PaddleOCR-VL:0.9b":
                # PaddleOCR-VL-1.5 使用chat API
                response = requests.post(
                    f"{ollama_url}/api/chat",
                    json={
                        "model": "my-PaddleOCR-VL:0.9b",
                        "messages": [
                            {
                                "role": "user",
                                "content": "请识别这张图片中的文字内容，并以Markdown格式输出。保留原有的段落结构和格式。",
                                "images": [img_base64]
                            }
                        ],
                        "stream": False,
                        "options": {
                            "temperature": 0.1,
                            "num_predict": 4096
                        }
                    },
                    timeout=300
                )
                
                if response.status_code == 200:
                    result = response.json()
                    content = result.get("message", {}).get("content", "")
                    if not content:
                        content = result.get("response", "")
                    if content:
                        all_content.append(f"## 第 {page_num} 页\n\n{content}")
                        st.success(f"✅ 第 {page_num}/{len(images)} 页处理完成")
                else:
                    st.error(f"❌ 第 {page_num}/{len(images)} 页识别失败: {response.status_code}")
            else:
                # GLM-OCR 使用Ollama原生 /api/generate 端点
                response = requests.post(
                    f"{ollama_url}/api/generate",
                    json={
                        "model": "my-glm-ocr:latest",
                        "prompt": "Text Recognition:",
                        "images": [img_base64],
                        "stream": False
                    },
                    timeout=300
                )
                
                if response.status_code == 200:
                    result = response.json()
                    content = result.get("response", "")
                    if not content:
                        content = result.get("message", {}).get("content", "")
                    if content:
                        all_content.append(f"## 第 {page_num} 页\n\n{content}")
                        st.success(f"✅ 第 {page_num}/{len(images)} 页处理完成")
                else:
                    st.error(f"❌ 第 {page_num}/{len(images)} 页识别失败: {response.status_code}")
        
        if all_content:
            markdown_content = "\n\n---\n\n".join(all_content)
            
            # 自动保存Markdown文件到docs文件夹
            try:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                base_name = os.path.splitext(filename)[0]
                md_filename = f"{base_name}_{timestamp}.md"
                md_filepath = os.path.join(DOCS_DIR, md_filename)
                
                with open(md_filepath, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                
                st.success(f"✅ Markdown文件已自动保存到: {md_filename}")
            except Exception as save_error:
                st.warning(f"⚠️ 保存Markdown文件失败: {str(save_error)}")
            
            return markdown_content
        else:
            st.warning("未能识别出任何内容")
            return ""
            
    except Exception as e:
        st.error(f"PDF转Markdown失败: {str(e)}")
        import traceback
        st.error(f"详细错误: {traceback.format_exc()}")
        return ""


def convert_doc_to_docx_bytes(doc_bytes: bytes) -> bytes:
    """使用Windows上的Word将doc转换为docx
    
    Args:
        doc_bytes: .doc文件的字节数据
        
    Returns:
        .docx文件的字节数据
    """
    import win32com.client as win32
    import pythoncom
    
    # 必须初始化COM，否则Streamlit多线程环境下会报错
    pythoncom.CoInitialize()
    
    # 将上传的字节流写入临时.doc文件
    with tempfile.NamedTemporaryFile(delete=False, suffix=".doc") as temp_doc:
        temp_doc.write(doc_bytes)
        temp_doc_path = temp_doc.name
        
    temp_docx_path = temp_doc_path + "x"  # 目标路径.docx
    
    try:
        word = win32.Dispatch("Word.Application")
        word.Visible = False
        # 打开.doc并另存为.docx (FileFormat=16表示docx格式)
        doc = word.Documents.Open(temp_doc_path)
        doc.SaveAs(temp_docx_path, FileFormat=16)
        doc.Close()
        word.Quit()
        
        # 读取生成的docx字节流
        with open(temp_docx_path, "rb") as f:
            docx_bytes = f.read()
        return docx_bytes
    finally:
        # 清理临时文件
        if os.path.exists(temp_doc_path):
            os.remove(temp_doc_path)
        if os.path.exists(temp_docx_path):
            os.remove(temp_docx_path)


def extract_text_from_docx(file_bytes: bytes) -> str:
    """从DOCX文件中提取文本"""
    try:
        from docx import Document
        doc = Document(io.BytesIO(file_bytes))
        text = []
        for para in doc.paragraphs:
            if para.text.strip():
                text.append(para.text)
        return "\n\n".join(text)
    except Exception as e:
        st.error(f"DOCX解析失败: {str(e)}")
        return ""


def extract_text_from_xlsx(file_bytes: bytes) -> str:
    """从XLSX文件中提取文本"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(file_bytes))
        text = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text.append(f"=== 工作表: {sheet_name} ===")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text.append(row_text)
        return "\n".join(text)
    except Exception as e:
        st.error(f"XLSX解析失败: {str(e)}")
        return ""


def batched(iterable, batch_size: int):
    batch = []
    for item in iterable:
        batch.append(item)
        if len(batch) >= batch_size:
            yield batch
            batch = []
    if batch:
        yield batch


def extract_text_from_image(image_bytes: bytes, ollama_url: str = "http://localhost:11434", model_name: str = "my-glm-ocr:latest", filename: str = "image.jpg") -> str:
    """使用Ollama的OCR模型从图像中提取文本
    
    Args:
        image_bytes: 图像文件的字节数据
        ollama_url: Ollama服务地址
        model_name: 使用的OCR模型名称
        filename: 原始文件名，用于保存Markdown文件
        
    Returns:
        识别出的文本内容
    """
    try:
        st.info(f"正在识别图像: {filename}")
        
        # 将图像转换为base64
        img_base64 = base64.b64encode(image_bytes).decode('utf-8')
        
        # 根据模型选择不同的调用方式
        if model_name == "my-PaddleOCR-VL:0.9b":
            # PaddleOCR-VL-1.5 使用chat API
            response = requests.post(
                f"{ollama_url}/api/chat",
                json={
                    "model": "my-PaddleOCR-VL:0.9b",
                    "messages": [
                        {
                            "role": "user",
                            "content": "请识别这张图片中的文字内容，并以Markdown格式输出。保留原有的段落结构和格式。",
                            "images": [img_base64]
                        }
                    ],
                    "stream": False,
                    "options": {
                        "temperature": 0.1,
                        "num_predict": 4096
                    }
                },
                timeout=300
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result.get("message", {}).get("content", "")
                if not content:
                    content = result.get("response", "")
                
                if content:
                    # 自动保存Markdown文件到docs文件夹
                    try:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        base_name = os.path.splitext(filename)[0]
                        md_filename = f"{base_name}_{timestamp}.md"
                        md_filepath = os.path.join(DOCS_DIR, md_filename)
                        
                        with open(md_filepath, 'w', encoding='utf-8') as f:
                            f.write(content)
                        
                        st.success(f"✅ 图像识别完成，Markdown文件已保存: {md_filename}")
                    except Exception as save_error:
                        st.warning(f"⚠️ 保存Markdown文件失败: {str(save_error)}")
                    
                    return content
                else:
                    st.warning("未能识别出任何内容")
                    return ""
            else:
                st.error(f"❌ 图像识别失败: {response.status_code}")
                return ""
        else:
            # GLM-OCR 使用Ollama原生 /api/generate 端点
            response = requests.post(
                f"{ollama_url}/api/generate",
                json={
                    "model": "my-glm-ocr:latest",
                    "prompt": "Text Recognition:",
                    "images": [img_base64],
                    "stream": False
                },
                timeout=300
            )
            
            if response.status_code == 200:
                result = response.json()
                content = result.get("response", "")
                if not content:
                    content = result.get("message", {}).get("content", "")
                
                if content:
                    # 自动保存Markdown文件到docs文件夹
                    try:
                        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                        base_name = os.path.splitext(filename)[0]
                        md_filename = f"{base_name}_{timestamp}.md"
                        md_filepath = os.path.join(DOCS_DIR, md_filename)
                        
                        with open(md_filepath, 'w', encoding='utf-8') as f:
                            f.write(content)
                        
                        st.success(f"✅ 图像识别完成，Markdown文件已保存: {md_filename}")
                    except Exception as save_error:
                        st.warning(f"⚠️ 保存Markdown文件失败: {str(save_error)}")
                    
                    return content
                else:
                    st.warning("未能识别出任何内容")
                    return ""
            else:
                st.error(f"❌ 图像识别失败: {response.status_code}")
                return ""
                
    except Exception as e:
        st.error(f"图像识别失败: {str(e)}")
        import traceback
        st.error(f"详细错误: {traceback.format_exc()}")
        return ""


def extract_text_from_file(uploaded_file, ollama_url: str = "http://localhost:11434", ocr_model: str = "my-glm-ocr:latest") -> str:
    """根据文件类型提取文本内容"""
    file_name = uploaded_file.name.lower()
    file_bytes = uploaded_file.read()
    
    if file_name.endswith('.txt') or file_name.endswith('.md'):
        return file_bytes.decode('utf-8')
    
    elif file_name.endswith('.doc'):
        # .doc文件先转换为.docx再提取文本
        st.info(f"正在将旧版Word文件 {uploaded_file.name} 转换为新版格式...")
        try:
            # 将doc转为docx的字节流
            docx_bytes = convert_doc_to_docx_bytes(file_bytes)
            # 复用原有的docx文本提取逻辑
            return extract_text_from_docx(docx_bytes)
        except Exception as e:
            st.error(f"DOC转换失败 (请确保系统已安装Microsoft Word): {str(e)}")
            return ""
    
    elif file_name.endswith('.docx'):
        return extract_text_from_docx(file_bytes)
    
    elif file_name.endswith('.xlsx'):
        return extract_text_from_xlsx(file_bytes)
    
    elif file_name.endswith('.pdf'):
        # PDF使用Ollama OCR转换为Markdown
        st.info(f"正在使用 {ocr_model} 处理PDF文件: {uploaded_file.name}")
        return pdf_to_markdown_with_ollama(file_bytes, ollama_url, ocr_model, uploaded_file.name)
    
    elif file_name.endswith(('.jpg', '.jpeg', '.png', '.bmp')):
        # 图像文件使用Ollama OCR识别
        st.info(f"正在使用 {ocr_model} 处理图像文件: {uploaded_file.name}")
        return extract_text_from_image(file_bytes, ollama_url, ocr_model, uploaded_file.name)
    
    else:
        st.warning(f"不支持的文件格式: {uploaded_file.name}")
        return ""


def process_uploaded_files(uploaded_files, index_name: str, namespace: str, ollama_url: str = "http://localhost:11434", ocr_model: str = "my-glm-ocr:latest") -> Tuple[int, int]:
    index = get_pinecone_index(index_name)
    if not index:
        return 0, 0

    total_docs = 0
    total_vectors = 0

    for uploaded_file in uploaded_files:
        try:
            content = extract_text_from_file(uploaded_file, ollama_url, ocr_model)
            if not content.strip():
                st.warning(f"文件 {uploaded_file.name} 内容为空，跳过")
                continue

            total_docs += 1
            chunks = split_text(content)

            vectors_to_upsert = []
            doc_id = total_docs

            for batch_id, batch_chunks in enumerate(batched(chunks, batch_size=32), start=1):
                embeddings = embed_texts(batch_chunks)

                for i, (chunk_text, embedding) in enumerate(zip(batch_chunks, embeddings), start=1):
                    vec_id = f"{uploaded_file.name}_b{batch_id}_c{i}"
                    vectors_to_upsert.append(
                        {
                            "id": vec_id,
                            "values": embedding,
                            "metadata": {
                                "source": uploaded_file.name,
                                "chunk_index": i,
                                "text": chunk_text,
                            },
                        }
                    )

            if vectors_to_upsert:
                index.upsert(vectors=vectors_to_upsert, namespace=namespace)
                total_vectors += len(vectors_to_upsert)
                st.success(f"✅ 文件 '{uploaded_file.name}' 处理完成，生成 {len(vectors_to_upsert)} 个向量")

        except Exception as e:
            st.error(f"处理文件 {uploaded_file.name} 时出错: {str(e)}")
            continue

    return total_docs, total_vectors


def pdf_to_markdown_page():
    """PDF转Markdown专用页面"""
    st.set_page_config(
        page_title="PDF转Markdown",
        page_icon="📄",
        layout="wide",
    )

    st.title("📄 PDF转Markdown工具")
    st.markdown("---")

    st.sidebar.title("⚙️ 配置")
    
    ollama_url = st.sidebar.text_input(
        "Ollama服务地址",
        value="http://localhost:11434",
        help="Ollama服务的URL地址"
    )
    
    ocr_model = st.sidebar.selectbox(
        "选择OCR模型",
        options=["my-PaddleOCR-VL:0.9b", "my-glm-ocr:latest"],
        index=1,
        help="选择用于PDF识别的OCR模型"
    )

    st.sidebar.info(
        f"""
        **使用说明:**
        
        1. 确保Ollama已安装并运行
        2. 确保已安装所选模型:
           `ollama pull {ocr_model}`
        3. 上传PDF文件
        4. 点击转换按钮
        5. 下载生成的Markdown文件
        
        **模型说明:**
        - **my-PaddleOCR-VL:0.9b**: PaddleOCR-VL-1.5，文档识别效果更好的模型
        - **my-glm-ocr:latest**: GLM-OCR，通用OCR模型
        """
    )

    uploaded_pdf = st.file_uploader(
        "选择PDF文件",
        type=["pdf"],
        help="上传需要转换为Markdown的PDF文件"
    )

    if uploaded_pdf:
        st.info(f"已选择文件: {uploaded_pdf.name}")
        st.info(f"使用模型: {ocr_model}")
        
        if st.button("🚀 开始转换", type="primary", use_container_width=True):
            with st.spinner(f"正在使用 {ocr_model} 模型转换PDF，请稍候..."):
                try:
                    pdf_bytes = uploaded_pdf.read()
                    markdown_content = pdf_to_markdown_with_ollama(pdf_bytes, ollama_url, ocr_model, uploaded_pdf.name)
                    
                    if markdown_content:
                        st.success("✅ 转换成功!")
                        
                        # 显示预览
                        with st.expander("预览Markdown内容"):
                            st.markdown(markdown_content[:2000] + "..." if len(markdown_content) > 2000 else markdown_content)
                        
                        # 提供下载
                        st.download_button(
                            label="📥 下载Markdown文件",
                            data=markdown_content.encode('utf-8'),
                            file_name=uploaded_pdf.name.replace('.pdf', '.md'),
                            mime="text/markdown"
                        )
                    else:
                        st.error("❌ 转换失败，请检查Ollama服务是否正常运行")
                        
                except Exception as e:
                    st.error(f"❌ 转换过程中出错: {str(e)}")


def document_upload_page():
    """文档上传主页面"""
    st.title("📄 文档上传到Pinecone")
    st.markdown("---")

    st.sidebar.title("⚙️ 配置")

    st.sidebar.info(
        """
        **设置说明:**
        
        1. 将 `.env.example` 复制为 `.env`
        2. 在 `.env` 中添加你的API密钥:
           - PINECONE_API_KEY
           - VECTORENGINE_API_KEY
        """
    )
    
    # Ollama配置
    st.sidebar.subheader("🤖 OCR配置")
    ollama_url = st.sidebar.text_input(
        "Ollama服务地址",
        value="http://localhost:11434",
        help="Ollama服务的URL地址，用于PDF识别"
    )
    
    ocr_model = st.sidebar.selectbox(
        "PDF识别模型",
        options=["my-PaddleOCR-VL:0.9b", "my-glm-ocr:latest"],
        index=1,
        help="选择用于PDF识别的OCR模型"
    )

    if not PINECONE_API_KEY or not VECTORENGINE_API_KEY:
        st.warning("⚠️ 请先在 `.env` 文件中配置API密钥!")
        st.stop()

    st.subheader("📚 索引管理")

    pc = Pinecone(api_key=PINECONE_API_KEY)
    existing_indexes = pc.list_indexes()
    index_names = [idx["name"] for idx in existing_indexes]

    col1, col2, col3 = st.columns([2, 1, 1])

    with col1:
        selected_index = st.selectbox(
            "选择索引",
            options=index_names if index_names else ["暂无索引"],
            index=0 if index_names else 0,
            key="index_selector",
            help="选择要使用的索引，如果没有索引请创建一个"
        )

    with col2:
        new_index_name = st.text_input(
            "新索引名称",
            placeholder="输入新索引名称",
            key="new_index_input",
            help="输入新索引名称后点击创建"
        )

    with col3:
        if st.button("🗑️ 删除索引", key="delete_index_button", use_container_width=True):
            if selected_index == "暂无索引":
                st.warning("没有可删除的索引!")
            else:
                try:
                    pc.delete_index(selected_index)
                    if "created_namespaces" in st.session_state and selected_index in st.session_state.created_namespaces:
                        del st.session_state.created_namespaces[selected_index]
                    st.success(f"✅ 索引 '{selected_index}' 已删除!")
                    st.rerun()
                except Exception as e:
                    st.error(f"删除索引失败: {str(e)}")

    col1, col2 = st.columns([1, 1])

    with col1:
        if st.button("➕ 创建新索引", key="create_index_button", use_container_width=True):
            if not new_index_name:
                st.warning("请输入索引名称!")
            elif new_index_name in index_names:
                st.warning("该索引已存在!")
            elif new_index_name != new_index_name.lower():
                st.warning("索引名称必须全部为小写字母!")
            elif not new_index_name.replace("-", "").isalnum():
                st.warning("索引名称只能包含小写字母、数字和连字符(-)!")
            else:
                try:
                    pc.create_index(
                        name=new_index_name,
                        dimension=3072,
                        metric="cosine",
                        spec=ServerlessSpec(
                            cloud=CloudProvider.AWS,
                            region=AwsRegion.US_EAST_1
                        ),
                    )
                    st.success(f"✅ 索引 '{new_index_name}' 创建成功!")
                    st.rerun()
                except Exception as e:
                    st.error(f"创建索引失败: {str(e)}")

    if not index_names:
        st.info("📌 暂无索引，请先创建一个索引!")
        st.stop()

    st.markdown("---")
    st.subheader(f"🏷️ 命名空间管理 (索引: {selected_index})")

    if selected_index and selected_index != "暂无索引":
        index = get_pinecone_index(selected_index)
        stats = index.describe_index_stats()
        existing_namespaces = list(stats.namespaces.keys()) if stats.namespaces else []
    else:
        existing_namespaces = []

    if "created_namespaces" not in st.session_state:
        st.session_state.created_namespaces = {}

    if selected_index not in st.session_state.created_namespaces:
        st.session_state.created_namespaces[selected_index] = []

    all_namespaces = list(set(existing_namespaces + st.session_state.created_namespaces[selected_index]))

    col1, col2, col3 = st.columns([2, 2, 1])

    with col1:
        selected_namespace = st.selectbox(
            "选择命名空间",
            options=["default"] + all_namespaces if all_namespaces else ["default"],
            index=0,
            key="namespace_selector",
            help=f"选择要使用的命名空间，default为默认命名空间（当前索引: {selected_index}）"
        )

    with col2:
        new_namespace = st.text_input(
            "新命名空间名称",
            placeholder="输入新命名空间名称",
            help="输入新命名空间名称后点击创建"
        )

    with col3:
        if st.button("➕ 创建", use_container_width=True):
            if not new_namespace:
                st.warning("请输入命名空间名称!")
            elif new_namespace in existing_namespaces or new_namespace == "default":
                st.warning("该命名空间已存在!")
            else:
                try:
                    temp_vector = [0.0] * 3072
                    temp_vector[0] = 1.0
                    index.upsert(
                        vectors=[{
                            "id": "temp_init_vector",
                            "values": temp_vector,
                            "metadata": {"temp": True}
                        }],
                        namespace=new_namespace
                    )
                    index.delete(ids=["temp_init_vector"], namespace=new_namespace)
                    st.session_state.created_namespaces[selected_index].append(new_namespace)
                    st.success(f"✅ 命名空间 '{new_namespace}' 创建成功!")
                    st.rerun()
                except Exception as e:
                    st.error(f"创建命名空间失败: {str(e)}")

    col1, col2 = st.columns([1, 1])

    with col1:
        if st.button("🗑️ 清空命名空间", use_container_width=True):
            if selected_namespace == "default":
                st.warning("不能清空默认命名空间!")
            else:
                try:
                    index.delete(delete_all=True, namespace=selected_namespace)
                    if selected_namespace in st.session_state.created_namespaces[selected_index]:
                        st.session_state.created_namespaces[selected_index].remove(selected_namespace)
                    st.success(f"✅ 命名空间 '{selected_namespace}' 已清空!")
                    st.rerun()
                except Exception as e:
                    st.error(f"清空命名空间失败: {str(e)}")

    st.markdown("---")
    st.subheader("上传文档")

    st.info("📌 支持两种上传方式：选择单个文件或选择文件夹进行批量上传")

    upload_mode = st.radio(
        "选择上传方式",
        ["📁 选择文件", "📂 选择文件夹"],
        horizontal=True,
        label_visibility="collapsed"
    )

    if upload_mode == "📁 选择文件":
        uploaded_files = st.file_uploader(
            "选择文件 (支持: txt, md, doc, docx, xlsx, pdf, jpg, jpeg, png, bmp)",
            type=["txt", "md", "doc", "docx", "xlsx", "pdf", "jpg", "jpeg", "png", "bmp"],
            accept_multiple_files=True,
            help="上传文本、Markdown、Word、Excel、PDF或图像文件到Pinecone"
        )
        folder_mode = False
    else:
        folder_path = st.text_input(
            "文件夹路径",
            placeholder="例如: D:\\Documents\\MyFiles",
            help="输入要批量上传的文件夹路径"
        )
        
        include_subfolders = st.checkbox(
            "包含子文件夹",
            value=True,
            help="是否递归处理子文件夹中的文件"
        )
        
        folder_files = []
        if folder_path:
            folder_files = get_files_from_folder(folder_path, include_subfolders)
            
            if folder_files:
                st.success(f"✅ 找到 {len(folder_files)} 个文件")
                
                with st.expander("📋 查看文件列表", expanded=False):
                    for i, file_path in enumerate(folder_files, 1):
                        rel_path = file_path.relative_to(folder_path) if folder_path else file_path
                        st.text(f"{i}. {rel_path} ({file_path.stat().st_size / 1024:.1f} KB)")
            else:
                st.warning("⚠️ 未找到支持的文件或文件夹路径无效")
        
        uploaded_files = folder_files
        folder_mode = True

    col1, col2 = st.columns([1, 1])

    with col1:
        max_chars = st.number_input(
            "每个文本块的最大字符数",
            min_value=100,
            max_value=5000,
            value=1000,
            step=100,
            help="每个文本块的最大字符数"
        )

    with col2:
        overlap = st.number_input(
            "重叠字符数",
            min_value=0,
            max_value=1000,
            value=100,
            step=50,
            help="文本块之间的重叠字符数"
        )

    if st.button("🚀 上传文档", type="primary", use_container_width=True):
        if not uploaded_files:
            st.warning("请至少上传一个文件或选择一个文件夹!")
            return

        with st.spinner("正在处理文档...这可能需要一些时间。"):
            try:
                if folder_mode:
                    progress_bar = st.progress(0, text="准备上传...")
                    total_docs = 0
                    total_vectors = 0
                    
                    for idx, file_path in enumerate(uploaded_files):
                        progress_bar.progress((idx + 1) / len(uploaded_files), text=f"正在处理 ({idx + 1}/{len(uploaded_files)}): {file_path.name}")
                        
                        try:
                            file_obj = create_file_object(file_path)
                            docs, vectors = process_uploaded_files([file_obj], selected_index, selected_namespace, ollama_url, ocr_model)
                            total_docs += docs
                            total_vectors += vectors
                        except Exception as e:
                            st.error(f"处理文件 {file_path.name} 时出错: {str(e)}")
                            continue
                    
                    progress_bar.empty()
                    
                    if selected_namespace != "default" and selected_namespace not in st.session_state.created_namespaces[selected_index]:
                        st.session_state.created_namespaces[selected_index].append(selected_namespace)
                    
                    st.success(f"✅ 上传完成!")
                    st.metric("已处理文档", total_docs)
                    st.metric("已创建向量", total_vectors)
                else:
                    total_docs, total_vectors = process_uploaded_files(uploaded_files, selected_index, selected_namespace, ollama_url, ocr_model)

                    if selected_namespace != "default" and selected_namespace not in st.session_state.created_namespaces[selected_index]:
                        st.session_state.created_namespaces[selected_index].append(selected_namespace)

                    st.success(f"✅ 上传完成!")
                    st.metric("已处理文档", total_docs)
                    st.metric("已创建向量", total_vectors)

            except Exception as e:
                st.error(f"❌ 上传过程中出错: {str(e)}")

    st.markdown("---")
    st.subheader(f"📊 索引状态: {selected_index}")

    try:
        index = get_pinecone_index(selected_index)
        if index:
            stats = index.describe_index_stats()
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.info(f"**当前命名空间**: `{selected_namespace}`")
                if stats.namespaces and selected_namespace in stats.namespaces:
                    ns_stats = stats.namespaces[selected_namespace]
                    st.metric(f"命名空间 '{selected_namespace}' 向量数量", ns_stats.vector_count)
                else:
                    st.metric(f"命名空间 '{selected_namespace}' 向量数量", 0)
            
            with col2:
                st.json({
                    "索引名称": selected_index,
                    "总向量数量": stats.total_vector_count,
                    "向量维度": stats.dimension,
                    "索引填充度": stats.index_fullness,
                    "所有命名空间": dict(stats.namespaces) if stats.namespaces else {}
                })
    except Exception as e:
        st.error(f"获取索引状态失败: {str(e)}")


def main():
    """主函数 - 页面导航"""
    st.sidebar.title("🧭 导航")
    
    page = st.sidebar.radio(
        "选择页面",
        ["📤 文档上传", "📄 PDF转Markdown"],
        help="选择要使用的功能页面"
    )
    
    if page == "📤 文档上传":
        document_upload_page()
    elif page == "📄 PDF转Markdown":
        pdf_to_markdown_page()


if __name__ == "__main__":
    main()
